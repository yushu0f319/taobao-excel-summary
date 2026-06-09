import importlib.util
import shutil
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "skills" / "taobao-daily-excel-update" / "scripts" / "update_taobao_daily.py"


def load_module():
    spec = importlib.util.spec_from_file_location("update_taobao_daily", SCRIPT)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def make_base_workbook(path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_headers = {
        "TB门店": [
            "日期",
            "总管id",
            "总管名称",
            "供应商id",
            "供应商名称",
            "城市名称",
            "门店名称",
            "门店id",
            "总成交额（元）",
            "预计收入（元）",
            "净营业额",
            "有效订单",
            "笔单价（元）",
            "顾客实付金额（元）",
        ],
        "TB流量": [
            "日期",
            "总管id",
            "总管名称",
            "供应商id",
            "供应商名称",
            "城市名称",
            "门店id",
            "门店名称",
            "总曝光人数",
            "总进店人数",
            "总下单人数",
            "进店转化率",
            "下单转化率",
        ],
        "TB履约": [
            "日期",
            "供应商id",
            "供应商名称",
            "门店名称",
            "门店id",
            "履约渗透率",
            "轨迹渗透率",
            "及时送达率",
            "流失订单",
            "流失订单率",
            "商家原因流失订单数",
            "商家原因流失订单率",
            "拣货及时订单数",
            "拣货及时订单率",
        ],
    }

    for sheet_name, headers in sheet_headers.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        if sheet_name == "TB门店":
            ws.append(
                [
                    20260604,
                    "17568075730400001",
                    "泉州大卖仓供应链管理有限公司",
                    "2000004631085",
                    "淘宝便利店（满佳喜）-戴总",
                    "镇江",
                    "淘宝便利店(句容店)",
                    1330837204,
                    100,
                    50,
                    80,
                    10,
                    10,
                    90,
                ]
            )
        elif sheet_name == "TB流量":
            ws.append(
                [
                    20260604,
                    "17568075730400001",
                    "泉州大卖仓供应链管理有限公司",
                    "2000004631085",
                    "淘宝便利店（满佳喜）-戴总",
                    "镇江",
                    1330837204,
                    "淘宝便利店(句容店)",
                    1000,
                    100,
                    10,
                    0.1,
                    0.1,
                ]
            )
            ws["L2"].number_format = "0.00%"
            ws["M2"].number_format = "0.00%"
        else:
            ws.append(
                [
                    20260604,
                    "2000004631085",
                    "淘宝便利店（满佳喜）-戴总",
                    "淘宝便利店(句容店)",
                    1330837204,
                    1,
                    1,
                    0.9,
                    2,
                    0.02,
                    1,
                    0.01,
                    90,
                    0.9,
                ]
            )
        ws.append([])
        ws.append(headers)
        # The final existing block supplies the style and type template.
        for col in range(1, len(headers) + 1):
            ws.cell(4, col)._style = ws.cell(1, col)._style
            ws.cell(5, col)._style = ws.cell(2, col)._style
        for col in range(1, len(headers) + 1):
            ws.cell(5, col).value = ws.cell(2, col).value
        ws.cell(5, 1).value = 20260604
        if sheet_name == "TB流量":
            ws["L5"].number_format = "0.00%"
            ws["M5"].number_format = "0.00%"
    wb.save(path)


def make_raw_workbook(path: Path, headers, data_row):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(headers)
    ws.append(data_row)
    meta = wb.create_sheet("meta")
    meta.append(["文件名", path.name])
    wb.save(path)

    # Simulate the Taobao export quirk where dimension says A1 despite real data.
    temp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(temp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(b'<dimension ref="A1:N2"/>', b'<dimension ref="A1"/>')
                data = data.replace(b'<dimension ref="A1:M2"/>', b'<dimension ref="A1"/>')
            zout.writestr(item, data)
    shutil.move(temp, path)


class UpdateTaobaoDailyTest(unittest.TestCase):
    def test_updates_three_taobao_sheets_and_converts_template_types(self):
        module = load_module()
        with tempfile.TemporaryDirectory() as temp_dir:
            tmp_path = Path(temp_dir)
            base = tmp_path / "6月4日度-规模数据.xlsx"
            output = tmp_path / "6月5日度-规模数据_AI复现.xlsx"
            make_base_workbook(base)

            make_raw_workbook(
                tmp_path / "门店.xlsx",
                [
                    "日期",
                    "总管id",
                    "总管名称",
                    "供应商id",
                    "供应商名称",
                    "城市名称",
                    "门店名称",
                    "门店id",
                    "总成交额（元）",
                    "预计收入（元）",
                    "净营业额",
                    "有效订单",
                    "笔单价（元）",
                    "顾客实付金额（元）",
                ],
                [
                    "20260605",
                    "17568075730400001",
                    "泉州大卖仓供应链管理有限公司",
                    "2000004631085",
                    "淘宝便利店（满佳喜）-戴总",
                    "镇江",
                    "淘宝便利店(句容店)",
                    "1330837204",
                    "15550.83",
                    "7093.68",
                    "10434.96",
                    "465",
                    "33.44",
                    "7348.24",
                ],
            )
            make_raw_workbook(
                tmp_path / "流量.xlsx",
                [
                    "日期",
                    "总管id",
                    "总管名称",
                    "供应商id",
                    "供应商名称",
                    "城市名称",
                    "门店id",
                    "门店名称",
                    "总曝光人数",
                    "总进店人数",
                    "总下单人数",
                    "进店转化率",
                    "下单转化率",
                ],
                [
                    "20260605",
                    "17568075730400001",
                    "泉州大卖仓供应链管理有限公司",
                    "2000004631085",
                    "淘宝便利店（满佳喜）-戴总",
                    "镇江",
                    "1330837204",
                    "淘宝便利店(句容店)",
                    "9439",
                    "1093",
                    "463",
                    "11.6%",
                    "42.4%",
                ],
            )
            make_raw_workbook(
                tmp_path / "履约.xlsx",
                [
                    "日期",
                    "供应商id",
                    "供应商名称",
                    "门店名称",
                    "门店id",
                    "履约渗透率",
                    "轨迹渗透率",
                    "及时送达率",
                    "流失订单",
                    "流失订单率",
                    "商家原因流失订单数",
                    "商家原因流失订单率",
                    "拣货及时订单数",
                    "拣货及时订单率",
                ],
                [
                    "20260605",
                    "2000004631085",
                    "淘宝便利店（满佳喜）-戴总",
                    "淘宝便利店(句容店)",
                    "1330837204",
                    "0.9978",
                    "0.9978",
                    "0.8923",
                    "12",
                    "0.0251",
                    "1",
                    "0.0021",
                    "378",
                    "0.8710",
                ],
            )

            report = module.update_workbook(
                base_path=base,
                raw_paths=[tmp_path / "门店.xlsx", tmp_path / "流量.xlsx", tmp_path / "履约.xlsx"],
                output_path=output,
            )

            self.assertEqual({item["kind"] for item in report["appended"]}, {"TB门店", "TB流量", "TB履约"})
            self.assertTrue(all(item["value_mismatches"] == 0 for item in report["appended"]))
            wb = openpyxl.load_workbook(output, data_only=False)
            self.assertEqual(wb["TB门店"]["A8"].value, 20260605)
            self.assertEqual(wb["TB门店"]["B8"].value, "17568075730400001")
            self.assertEqual(wb["TB门店"]["H8"].value, 1330837204)
            self.assertEqual(wb["TB门店"]["I8"].value, 15550.83)
            self.assertEqual(wb["TB流量"]["G8"].value, 1330837204)
            self.assertEqual(wb["TB流量"]["L8"].value, 0.116)
            self.assertEqual(wb["TB流量"]["L8"].number_format, "0.00%")
            self.assertEqual(wb["TB履约"]["E8"].value, 1330837204)
            self.assertEqual(wb["TB履约"]["F8"].value, 0.9978)
            wb.close()


if __name__ == "__main__":
    unittest.main()
